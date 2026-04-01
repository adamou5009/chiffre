# extractionAlonwaOptimise.py
import streamlit as st
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
import time
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from io import BytesIO
from datetime import datetime

from config_driver import WebDriverManager, TIMEOUT

# -------------------------------
# Récupération des paramètres
# -------------------------------
username = st.session_state.get('username')
password = st.session_state.get('password')
statuts_choisis = st.session_state.get('statuts_choisis')
date_debut = st.session_state.get('date_debut').strftime("%d/%m/%Y")
date_fin = st.session_state.get('date_fin').strftime("%d/%m/%Y")

url = "https://serviceplus.canal-plus.com/index.php?action=GET_LOGIN"  # Remplacer par l'URL réelle

# -------------------------------
# Fonction de connexion
# -------------------------------
def connexion_alonwa(driver_instance, url, username, password, timeout=TIMEOUT):
    wait = WebDriverWait(driver_instance, timeout)
    try:
        driver_instance.get(url)
        champ_identifiant = wait.until(EC.element_to_be_clickable((By.ID, "in_username")))
        champ_mot_de_passe = wait.until(EC.element_to_be_clickable((By.ID, "in_password")))
        champ_identifiant.send_keys(username)
        champ_mot_de_passe.send_keys(password)
        bouton_connexion = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//input[@type='submit' and contains(@class, 'newimgbtn')]"))
        )
        bouton_connexion.click()
        wait.until(EC.presence_of_element_located((By.ID, "divContainer")))
        return True, driver_instance, wait
    except (TimeoutException, NoSuchElementException, WebDriverException) as e:
        st.error(f"❌ Erreur lors de la connexion : {e}")
        return False, None, None

# -------------------------------
# Sélection de date
# -------------------------------
def selectionner_date(driver, wait, champ_id, date_str):
    jour, mois, annee = map(int, date_str.split("/"))
    mois -= 1
    champ = wait.until(EC.element_to_be_clickable((By.ID, champ_id))).click()
    driver.execute_script("arguments[0].click();", champ)
    time.sleep(0.5)
    MOIS_TEXT = ["Janvier","Février","Mars","Avril","Mai","Juin","Juillet",
                 "Août","Septembre","Octobre","Novembre","Décembre"]
    while True:
        titre = driver.find_element(By.CLASS_NAME, "ui-datepicker-title")
        mois_affiche = titre.find_element(By.CLASS_NAME, "ui-datepicker-month").text
        annee_affiche = int(titre.find_element(By.CLASS_NAME, "ui-datepicker-year").text)
        mois_actuel = MOIS_TEXT.index(mois_affiche)
        if annee_affiche < annee or (annee_affiche == annee and mois_actuel < mois):
            driver.find_element(By.CLASS_NAME, "ui-icon-circle-triangle-e").click()
        elif annee_affiche > annee or (annee_affiche == annee and mois_actuel > mois):
            driver.find_element(By.CLASS_NAME, "ui-icon-circle-triangle-w").click()
        else:
            break
        time.sleep(0.2)
    xpath_jour = f"//table[@class='ui-datepicker-calendar']//a[text()='{jour}']"
    jour_element = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_jour)))
    jour_element.click()
    time.sleep(0.3)

# -------------------------------
# Sélection statuts multi
# -------------------------------
def selectionner_statuts(driver, wait, statuts_choisis):
    DROPDOWN_ID = "intervention_status_select-button"
    MENU_ID = "intervention_status_select-menu"
    dropdown = wait.until(EC.element_to_be_clickable((By.ID, DROPDOWN_ID)))
    driver.execute_script("arguments[0].click();", dropdown)
    time.sleep(0.3)
    for statut in statuts_choisis:
        xpath_option = f"//ul[@id='{MENU_ID}']/li[contains(text(),'{statut}')]"
        option = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_option)))
        driver.execute_script("arguments[0].click();", option)
        time.sleep(0.2)

# -------------------------------
# Extraction tableau
# -------------------------------
def extraire_tableau(driver, wait, statuts_choisis):
    resultats = []
    time.sleep(1)
    try:
        driver.find_element(By.ID, "tbl_inter_pending_first").click()
        time.sleep(1)
    except: pass

    while True:
        lignes = driver.find_elements(
            By.XPATH,
            "//table[@id='tbl_inter_pending']/tbody/tr[not(contains(@class,'dataTables_empty'))]"
        )
        for tr in lignes:
            colonnes = tr.find_elements(By.TAG_NAME, "td")
            if len(colonnes) >= 8:
                numero = colonnes[4].text.strip()
                for statut_choisi in statuts_choisis:
                    resultats.append((numero, statut_choisi))
        try:
            next_btn = driver.find_element(By.ID, "tbl_inter_pending_next")
            if "ui-state-disabled" in next_btn.get_attribute("class"):
                break
            next_btn.click()
            time.sleep(1)
        except:
            break
    return resultats

# -------------------------------
# Génération Excel en mémoire
# -------------------------------
def generer_excel_en_memoire(resultats):
    wb = Workbook()
    ws = wb.active
    ws.title = "Abonnés"
    ws.append(["Numero_Abonne", "Statut_Trouve"])
    vert_fill = PatternFill(start_color="00C6EFCE", end_color="00C6EFCE", fill_type="solid")
    for numero, statut in resultats:
        ws.append([numero, statut])
        if statut == "Terminée OK":
            for cell in ws[ws.max_row]:
                cell.fill = vert_fill
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# -------------------------------
# Lancement
# -------------------------------
if st.button("📥 Lancer l'extraction"):
    driver = WebDriverManager().get_driver()
    ok, driver, wait = connexion_alonwa(driver, url, username, password)
    if ok:
        st.info("Connexion réussie ! Application des filtres...")
        selectionner_date(driver, wait, "intervention_from_datecrea", date_debut)
        selectionner_date(driver, wait, "intervention_to_datecrea", date_fin)
        selectionner_statuts(driver, wait, statuts_choisis)
        st.info("Extraction des données en cours...")
        resultats = extraire_tableau(driver, wait, statuts_choisis)
        fichier_excel = generer_excel_en_memoire(resultats)
        st.success(f"✅ Extraction terminée !")
        st.download_button(
            "Télécharger le fichier Excel",
            data=fichier_excel,
            file_name="resultats_abonnes.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.error("❌ Connexion impossible, vérifie les identifiants.")
