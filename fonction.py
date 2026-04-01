import time
import re   
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException, StaleElementReferenceException

# La constante TIMEOUT est supposée être importée de config_driver
# Si non définie, utilisez une valeur par défaut, ex: TIMEOUT = 15

# -------------------------------
# Connexion à l'application SAVANT
# -------------------------------
def connexion_savant(driver_instance, url, username, password, timeout=15):
    """Connexion à l'application SAVANT"""
    wait = WebDriverWait(driver_instance, timeout)
    try:
        driver_instance.get(url)
        # Attente des champs de connexion
        champ_identifiant = wait.until(EC.element_to_be_clickable((By.ID, "in_username")))
        champ_mot_de_passe = wait.until(EC.element_to_be_clickable((By.ID, "in_password")))
        
        champ_identifiant.send_keys(username)
        champ_mot_de_passe.send_keys(password)
        
        # Attente et clic sur le bouton de connexion
        bouton_connexion = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//input[@type='submit' and contains(@class, 'newimgbtn')]"))
        )
        bouton_connexion.click()
        
        # ✅ Attente d'un élément post-connexion fiable (ex: la barre de menu)
        wait.until(EC.presence_of_element_located((By.ID, "menudiv"))) 
        print("✅ Connexion réussie.")
        return True, driver_instance, wait
        
    except (TimeoutException, NoSuchElementException, WebDriverException) as e:
        print(f"❌ Erreur connexion ({e.__class__.__name__}): {e}")
        return False, None, None

# -------------------------------
# Navigation vers la page Intervention
# -------------------------------
def naviguer_page_intervention(driver_instance, wait):
    """Clique sur le lien 'Intervention' pour naviguer vers la page de liste."""
    try:
        # Assurez-vous d'être hors de tout iframe
        driver_instance.switch_to.default_content() 
        
        wait.until(EC.presence_of_element_located((By.ID, "menudiv"))) 
        intervention_btn = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//a[contains(@href, 'INTER_PENDING')]"))
        )
        driver_instance.execute_script("arguments[0].click();", intervention_btn)
        
        # Attendre qu'un élément de la nouvelle page soit présent
        wait.until(EC.presence_of_element_located((By.ID, "home_body"))) 
        print("✅ Page Intervention chargée.")
        return True
        
    except (TimeoutException, StaleElementReferenceException) as e:
        print(f"❌ Erreur lors du clic sur 'intervention' : {e.__class__.__name__}")
        return False

# -------------------------------
# Sélection de dates dans le datepicker
# -------------------------------
def selectionner_date_avec_validation(driver, wait, champ_id, date_str, bouton_valider_id="btn_period_valid"):
    """Sélectionne une date dans le datepicker et valide le filtre."""
    try:
        jour, mois, annee = map(int, date_str.split("/"))
        mois_index = mois - 1  # JS datepicker 0-indexé (0=Janvier)
    except ValueError:
        print(f"❌ Format de date incorrect: {date_str}. Utiliser JJ/MM/AAAA.")
        return False
        
    MOIS_TEXT = ["Janvier","Février","Mars","Avril","Mai","Juin",
                 "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]

    # 1️⃣ Cliquer sur le champ
    try:
        champ = wait.until(EC.element_to_be_clickable((By.ID, champ_id)))
        driver.execute_script("arguments[0].click();", champ)
        time.sleep(0.5)
    except TimeoutException:
        print(f"❌ Champ de date non cliquable: {champ_id}")
        return False

    # 2️⃣ Attendre que le datepicker soit visible
    try:
        wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "ui-datepicker-title")))
    except TimeoutException:
        print("❌ Datepicker non visible après clic.")
        return False

    # 3️⃣ Naviguer mois par mois
    while True:
        try:
            nav_prev = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "ui-icon-circle-triangle-w")))
            nav_next = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "ui-icon-circle-triangle-e")))
            titre = driver.find_element(By.CLASS_NAME, "ui-datepicker-title")
            
            mois_affiche = titre.find_element(By.CLASS_NAME, "ui-datepicker-month").text
            annee_affiche = int(titre.find_element(By.CLASS_NAME, "ui-datepicker-year").text)
            mois_actuel = MOIS_TEXT.index(mois_affiche)
        except (NoSuchElementException, StaleElementReferenceException, ValueError):
            print("❌ Erreur lors de la lecture du mois/année affiché pendant la navigation.")
            return False

        if annee_affiche < annee or (annee_affiche == annee and mois_actuel < mois_index):
            nav_next.click() 
        elif annee_affiche > annee or (annee_affiche == annee and mois_actuel > mois_index):
            nav_prev.click() 
        else:
            break
        time.sleep(0.2)

    # 4️⃣ Cliquer sur le jour
    try:
        xpath_jour = f"//table[@class='ui-datepicker-calendar']//a[text()='{jour}']"
        jour_element = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_jour)))
        jour_element.click()
        time.sleep(0.3)
    except TimeoutException:
        print(f"❌ Jour {jour} non cliquable ou introuvable.")
        return False

    # 5️⃣ Cliquer sur le bouton Valider
    try:
        bouton_valider = wait.until(EC.element_to_be_clickable((By.ID, bouton_valider_id)))
        driver.execute_script("arguments[0].click();", bouton_valider)
        time.sleep(1)
        print(f"✅ Date {date_str} sélectionnée et filtre validé.")
    except TimeoutException:
        print(f"❌ Bouton Valider non cliquable: {bouton_valider_id}")
        return False

    return True

#correction de la fonction
def activer_et_selectionner_dates(driver, wait, date_debut, date_fin,
                                  champ_debut_id="intervention_from_datecrea",
                                  champ_fin_id="intervention_to_datecrea",
                                  div_dates_id="intervention_date_div",
                                  bouton_valider_id="btn_period_valid"):
    """
    Active le sélecteur de dates et configure la période complète.
    Cliquer sur chaque champ avant de sélectionner la date.
    """
    # ✅ 1️⃣ Cliquer sur le champ début pour ouvrir le calendrier
    try:
        champ_from = wait.until(EC.element_to_be_clickable((By.ID, champ_debut_id)))
        #driver.execute_script("arguments[0].click();", champ_from)
        champ_from.click()
        wait.until(EC.visibility_of_element_located((By.ID, div_dates_id)))
        print("✅ Calendrier visible pour date début")
    except TimeoutException:
        print("❌ Impossible d'activer le champ date début")
        return False

    # ✅ 2️⃣ Sélectionner la date début sans valider
    if not selectionner_date_avec_validation(driver, wait, champ_debut_id, date_debut, valider=False):
        print("❌ Sélection date début échouée")
        return False

    # ✅ 3️⃣ Cliquer sur le champ fin pour ouvrir le calendrier
    try:
        champ_to = wait.until(EC.element_to_be_clickable((By.ID, champ_fin_id)))
        #driver.execute_script("arguments[0].click();", champ_to)
        champ_to.click()
        time.sleep(0.3)
        print("✅ Calendrier visible pour date fin")
    except TimeoutException:
        print("❌ Impossible d'activer le champ date fin")
        return False

    # ✅ 4️⃣ Sélectionner la date fin sans valider
    if not selectionner_date_avec_validation(driver, wait, champ_fin_id, date_fin, valider=False):
        print("❌ Sélection date fin échouée")
        return False

    # ✅ 5️⃣ Valider les deux dates
    try:
        bouton_valider = wait.until(EC.element_to_be_clickable((By.ID, bouton_valider_id)))
        driver.execute_script("arguments[0].click();", bouton_valider)
        time.sleep(1)
        print(f"✅ Période validée: {date_debut} → {date_fin}")
        return True
    except TimeoutException:
        print("❌ Bouton Valider non cliquable")
        return False



def selectionner_date_avec_validation(driver, wait, champ_id, date_str, valider=True):
    """
    Sélectionne une date dans le datepicker.
    
    Args:
        valider: Si False, ne clique pas sur le bouton Valider (pour valider les 2 dates ensemble)
    """
    # il faut cliquer sur #intervention_from_datecrea pour ouvir le calendrier

    try:
        jour, mois, annee = map(int, date_str.split("/"))
        mois_index = mois - 1
    except ValueError:
        print(f"❌ Format de date incorrect: {date_str}")
        return False
        
    MOIS_TEXT = ["Janvier","Février","Mars","Avril","Mai","Juin",
                 "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]

    # 1️⃣ Cliquer sur le champ pour ouvrir le datepicker
    try:
        champ = wait.until(EC.element_to_be_clickable((By.ID, champ_id)))
        driver.execute_script("arguments[0].click();", champ)
        time.sleep(0.5)
    except TimeoutException:
        print(f"❌ Champ {champ_id} non cliquable")
        return False

    # 2️⃣ Attendre que le datepicker soit visible
    try:
        wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "ui-datepicker-title")))
    except TimeoutException:
        print("❌ Datepicker non visible")
        return False

    # 3️⃣ Naviguer vers le bon mois/année
    max_iterations = 24
    iteration = 0
    
    while iteration < max_iterations:
        try:
            nav_prev = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "ui-icon-circle-triangle-w")))
            nav_next = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "ui-icon-circle-triangle-e")))
            titre = driver.find_element(By.CLASS_NAME, "ui-datepicker-title")
            
            mois_affiche = titre.find_element(By.CLASS_NAME, "ui-datepicker-month").text
            annee_affiche = int(titre.find_element(By.CLASS_NAME, "ui-datepicker-year").text)
            mois_actuel = MOIS_TEXT.index(mois_affiche)
            
        except (NoSuchElementException, StaleElementReferenceException, ValueError) as e:
            print(f"❌ Erreur navigation: {e}")
            return False

        # Navigation
        if annee_affiche < annee or (annee_affiche == annee and mois_actuel < mois_index):
            nav_next.click()
        elif annee_affiche > annee or (annee_affiche == annee and mois_actuel > mois_index):
            nav_prev.click()
        else:
            break
            
        time.sleep(0.2)
        iteration += 1

    if iteration >= max_iterations:
        print("❌ Limite de navigation atteinte")
        return False

    # 4️⃣ Cliquer sur le jour
    try:
        xpath_jour = f"//table[@class='ui-datepicker-calendar']//a[text()='{jour}']"
        jour_element = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_jour)))
        jour_element.click()
        time.sleep(0.3)
        
        # Vérifier que la valeur est mise à jour
        champ_updated = driver.find_element(By.ID, champ_id)
        valeur = champ_updated.get_attribute("value")
        print(f"  ✓ Date sélectionnée: {valeur}")
        
    except TimeoutException:
        print(f"❌ Jour {jour} non trouvé")
        return False

    # 5️⃣ Valider si demandé (optionnel)
    if valider:
        try:
            bouton_valider = wait.until(EC.element_to_be_clickable((By.ID, "btn_period_valid")))
            driver.execute_script("arguments[0].click();", bouton_valider)
            time.sleep(1)
            print(f"✅ Date {date_str} validée")
        except TimeoutException:
            print("❌ Bouton Valider non cliquable")
            return False

    return True



# -------------------------------
# Sélection des statuts (multi)
# -------------------------------
def selectionner_statuts(driver, wait, statuts_choisis):
    """
    Sélectionne les statuts dans le dropdown 'Intervention Status'.
    """
    DROPDOWN_ID = "intervention_status_select-button"
    MENU_ID = "intervention_status_select-menu"
    VALIDER_ID = "btn_period_valid" 

    def cliquer_option(texte_option):
        """Clique sur une option avec retry pour StaleElementReference"""
        for _ in range(3):
            try:
                xpath_option = f"//ul[@id='{MENU_ID}']/li[contains(normalize-space(.), '{texte_option}')]"
                option = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_option)))
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", option)
                option.click() 
                return True
            except StaleElementReferenceException:
                time.sleep(0.2)
            except TimeoutException:
                 print(f"⚠ Option introuvable ou non cliquable: {texte_option}")
                 return False
        return False

    # 1️⃣ Ouvrir le dropdown
    try:
        dropdown = wait.until(EC.element_to_be_clickable((By.ID, DROPDOWN_ID)))
        driver.execute_script("arguments[0].click();", dropdown)
        time.sleep(0.3)
    except TimeoutException:
        print(f"❌ Impossible d'ouvrir le dropdown : {DROPDOWN_ID}")
        return

    # 2️⃣ Cliquer sur chaque statut choisi
    for statut in statuts_choisis:
        cliquer_option(statut)
        time.sleep(0.2)

    # 3️⃣ Valider le filtre
    try:
        bouton_valider = wait.until(EC.element_to_be_clickable((By.ID, VALIDER_ID)))
        driver.execute_script("arguments[0].click();", bouton_valider)
        time.sleep(1)
        print("✅ Statuts sélectionnés et filtre validé.")
    except TimeoutException:
        print(f"❌ Bouton Valider non cliquable: {VALIDER_ID}")
        
# -------------------------------
# Extraction du tableau
# -------------------------------
# NOTE: Cette fonction est conservée ici car elle sera appelée par PageExtraction.py
# fonction.py

# ... (Début du fichier avec les imports)

# -------------------------------
# Extraction du tableau
# -------------------------------
def extraire_tableau(driver, wait, statuts_choisis):
    """Extraction complète et fiable de toutes les pages."""
    resultats = []
    numeros_vus = set()

    # ── Aller à la première page ──────────────────────────
    try:
        wait.until(EC.element_to_be_clickable((By.ID, "tbl_inter_pending_first"))).click()
        time.sleep(2)
    except Exception:
        pass

    page_count = 0

    while True:
        if arret_demande():
            print("⛔ Arrêt demandé")
            break

        page_count += 1

        # ── Attendre que le tableau soit prêt ─────────────
        try:
            wait.until(EC.invisibility_of_element_located(
                (By.ID, "tbl_inter_pending_processing")
            ))
        except Exception:
            pass

        try:
            wait.until(EC.presence_of_element_located(
                (By.XPATH, "//table[@id='tbl_inter_pending']/tbody/tr")
            ))
        except Exception:
            pass

        time.sleep(0.5)

        # ── Extraire TOUTES les lignes de la page ─────────
        lignes = driver.find_elements(
            By.XPATH,
            "//table[@id='tbl_inter_pending']/tbody/tr"
            "[not(contains(@class,'dataTables_empty'))]"
        )

        nb_lignes = len(lignes)
        print(f"📄 Page {page_count} — {nb_lignes} lignes")

        if nb_lignes == 0:
            print("  → Tableau vide, arrêt")
            break

        # ── Lire chaque ligne avec retry anti-stale ───────
        for idx in range(nb_lignes):
            if arret_demande():
                break

            for tentative in range(3):
                try:
                    lignes_actuelles = driver.find_elements(
                        By.XPATH,
                        "//table[@id='tbl_inter_pending']/tbody/tr"
                        "[not(contains(@class,'dataTables_empty'))]"
                    )

                    if idx >= len(lignes_actuelles):
                        break

                    tr = lignes_actuelles[idx]
                    colonnes = tr.find_elements(By.TAG_NAME, "td")

                    if len(colonnes) < 5:
                        break

                    numero = colonnes[4].text.strip()

                    if not numero:
                        break

                    if numero not in numeros_vus:
                        numeros_vus.add(numero)
                        resultats.append({
                            "Numero": numero,
                            "Statut": ", ".join(statuts_choisis)
                        })

                    break

                except StaleElementReferenceException:
                    print(f"  ⚠️ Stale ligne {idx}, tentative {tentative + 1}/3")
                    time.sleep(0.3)
                    continue
                except Exception as e:
                    print(f"  ⚠️ Erreur ligne {idx} : {e}")
                    break

        print(f"  → Cumulé : {len(resultats)} interventions extraites")

        # ── Pagination sans référence à un WebElement ─────
        try:
            # ✅ Capturer le texte AVANT le clic (string pure, pas WebElement)
            lignes_avant = driver.find_elements(
                By.XPATH,
                "//table[@id='tbl_inter_pending']/tbody/tr"
                "[not(contains(@class,'dataTables_empty'))]"
            )
            # ✅ Extraire le texte immédiatement en string pure
            texte_avant = lignes_avant[0].find_elements(
                By.TAG_NAME, "td"
            )[4].text.strip() if lignes_avant else ""

            # Vérifier si Next est disponible
            next_btn = driver.find_element(By.ID, "tbl_inter_pending_next")
            classes = next_btn.get_attribute("class") or ""

            if "disabled" in classes or "ui-state-disabled" in classes:
                print("✅ Dernière page atteinte")
                break

            # Clic JS
            driver.execute_script("arguments[0].click();", next_btn)

            # ✅ Attendre en comparant uniquement des strings (pas de WebElement)
            for _ in range(30):  # max 15 secondes
                time.sleep(0.5)
                try:
                    nouvelles_lignes = driver.find_elements(
                        By.XPATH,
                        "//table[@id='tbl_inter_pending']/tbody/tr"
                        "[not(contains(@class,'dataTables_empty'))]"
                    )
                    if not nouvelles_lignes:
                        continue
                    texte_apres = nouvelles_lignes[0].find_elements(
                        By.TAG_NAME, "td"
                    )[4].text.strip()

                    if texte_apres != texte_avant:
                        break  # ✅ page changée
                except Exception:
                    continue

        except Exception as e:
            print(f"  → Fin pagination ({e})")
            break

    print(f"\n🏁 TOTAL : {len(resultats)} interventions — Statuts : {statuts_choisis}")
    return resultats



       

# -------------------------------
# Génération du fichier Excel en mémoire
# -------------------------------
def generer_excel_en_memoire(resultats):
    """Crée un fichier Excel à partir des résultats et le retourne sous forme de buffer en mémoire."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Abonnés"
    ws.append(["Numero_Abonne", "Statut_Trouve"])
    
    # Remplissage vert pour les statuts "Terminée OK"
    vert_fill = PatternFill(start_color="00C6EFCE", end_color="00C6EFCE", fill_type="solid")
    
    for numero, statut in resultats:
        ws.append([numero, statut])
        if statut == "Terminée OK":
            # Applique le remplissage à toutes les cellules de la ligne
            for cell in ws[ws.max_row]:
                cell.fill = vert_fill
                
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# -------------------------------
#fonction pour selectionner les temporaires
# -------------------------------
def selectionner_statut_temporaire(driver, wait):
    try:
        dropdown = wait.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "#intervention_status_select-button"))
        )
        dropdown.click()

        option_temp = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//li[contains(., 'Temporaire')]"))
        )
        option_temp.click()

        wait.until(EC.presence_of_element_located((By.ID, "tbl_inter_pending")))
        return True

    except Exception as e:
        print("Erreur sélection statut Temporaire :", e)
        return False

# -------------------------------
# Gestion arrêt du traitement
# -------------------------------
STOP_REQUESTED = False

def demander_arret():
    global STOP_REQUESTED
    STOP_REQUESTED = True

def reset_arret():
    global STOP_REQUESTED
    STOP_REQUESTED = False

def arret_demande():
    return STOP_REQUESTED



# -------------------------------
#fonction pour extraire les temporaires
# -------------------------------
def extraire_interventions_temporaire(driver, wait):
    resultats = []
    decodeurs_vus = set()
    page_num = 1

    while True:

        if arret_demande():
            print("⛔ Arrêt demandé – retour résultats temporaires")
            break

        wait.until(EC.presence_of_element_located((By.ID, "tbl_inter_pending")))
        voir_btns = driver.find_elements(By.XPATH, "//a[contains(text(),'Voir')]")

        for i, btn in enumerate(voir_btns, start=1):

            if arret_demande():
                break

            try:
                driver.execute_script("arguments[0].click();", btn)
                wait.until(lambda d: len(d.window_handles) > 1)
                driver.switch_to.window(driver.window_handles[-1])

                wait.until(EC.presence_of_all_elements_located(
                    (By.CSS_SELECTOR, "h3.ui-accordion-header"))
                )

                headers = driver.find_elements(By.CSS_SELECTOR, "h3.ui-accordion-header")

                header_rdv = header_cr = None
                for h in headers:
                    t = h.text.lower()
                    if "rdv" in t:
                        header_rdv = h
                    if "compte" in t or "rendu" in t:
                        header_cr = h

                id_tech = None
                if header_rdv:
                    header_rdv.click()
                    time.sleep(0.5)
                    try:
                        el = driver.find_element(
                            By.XPATH,
                            "//*[contains(text(),'ID Tech')]/following::div[1]"
                        )
                        id_tech = el.text.strip()
                    except:
                        pass
                    header_rdv.click()

                decodeurs = []
                if header_cr:
                    header_cr.click()
                    time.sleep(0.5)
                    inputs = driver.find_elements(
                        By.XPATH, "//input[starts-with(@id,'ref_decodeur')]")
                    for inp in inputs:
                        val = inp.get_attribute("value").strip()
                        if re.fullmatch(r"\d{14}", val) and val not in decodeurs_vus:
                            decodeurs.append(val)
                            decodeurs_vus.add(val)
                    header_cr.click()

                if decodeurs:
                    resultats.append({
                        "Page": page_num,
                        "Ligne": i,
                        "ID Tech": id_tech,
                        "Décodeurs": ", ".join(decodeurs)
                    })

                driver.close()
                driver.switch_to.window(driver.window_handles[0])

            except Exception as e:
                print("⚠️ Erreur ligne temporaire :", e)

        try:
            next_btn = driver.find_element(By.XPATH, "//a[contains(text(),'Suivant')]")
            if "ui-state-disabled" in next_btn.get_attribute("class"):
                break
            next_btn.click()
            page_num += 1
            time.sleep(1.5)
        except:
            break

    return resultats

# -------------------------------
# EXPORT EXCEL MULTI-FEUILLES
#--------------------------------
import io
from openpyxl import Workbook

def generer_excel_multi_feuilles(data_temporaire, data_autres):
    """
    Génère un fichier Excel avec deux feuilles :
    - TEMPORAIRE
    - AUTRES STATUTS
    """
    wb = Workbook()
    
    # ================================
    # Feuille TEMPORAIRE
    # ================================
    ws_t = wb.active
    ws_t.title = "TEMPORAIRE"
    
    if data_temporaire:
        # En-têtes
        ws_t.append(list(data_temporaire[0].keys()))
        # Lignes
        for row in data_temporaire:
            ws_t.append(list(row.values()))
    else:
        ws_t.append(["Aucune donnée TEMPORAIRE"])

    # ================================
    # Feuille AUTRES STATUTS
    # ================================
    ws_a = wb.create_sheet(title="AUTRES STATUTS")
    
    if data_autres:
        # En-têtes
        ws_a.append(list(data_autres[0].keys()))
        # Lignes
        for row in data_autres:
            ws_a.append(list(row.values()))
    else:
        ws_a.append(["Aucune donnée AUTRES STATUTS"])

    # ================================
    # Sauvegarde en mémoire
    # ================================
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

# -------------------------------
# Vérification si "Temporaire" est dans les statuts sélectionnés
# -------------------------------
def est_statut_temporaire(statuts):
    return "Temporaire" in statuts


def selectionner_statuts_pipeline(driver, wait, statuts):
    if statuts == ["Temporaire"]:
        return selectionner_statut_temporaire(driver, wait)
    else:
        return selectionner_statuts(driver, wait, statuts)


