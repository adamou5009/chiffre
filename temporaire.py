from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.options import Options
from selenium.common.exceptions import NoSuchElementException
import time
import re
import os
from openpyxl import Workbook, load_workbook

# --- PARAMÈTRES ---
URL_LOGIN = "https://serviceplus.canal-plus.com/index.php?action=INDEX"
LOGIN = "Centre_Rayinda"          # <<< ton identifiant
PASSWORD = "7LRvuZlqitVMNI4u"       # <<< ton mot de passe

# Sélecteurs CSS
CSS_ID_INPUT = "#in_username"
CSS_PW_INPUT = "#in_password"
CSS_LOGIN_BUTTON = "#divLogin > form > div.login_footer_div > input"

# --- CONFIG EDGE ---
options = Options()
options.add_argument("--start-maximized")
driver = webdriver.Edge(options=options)
wait = WebDriverWait(driver, 20)

# --- 1. OUVERTURE PAGE LOGIN ---
driver.get(URL_LOGIN)

# --- 2. CONNEXION ---
id_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, CSS_ID_INPUT)))
id_input.send_keys(LOGIN)

pw_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, CSS_PW_INPUT)))
pw_input.send_keys(PASSWORD)

login_btn = driver.find_element(By.CSS_SELECTOR, CSS_LOGIN_BUTTON)
login_btn.click()

# --- 3. CLIC SUR LE LIEN "Intervention" ---
intervention_link = wait.until(
    EC.element_to_be_clickable((By.XPATH, '//a[@href="https://serviceplus.canal-plus.com/index.php?action=INTER_PENDING"]'))
)
intervention_link.click()
time.sleep(2)

# --- 4. CHOIX DU STATUT "Temporaire" ---
CSS_DROPDOWN_BTN = "#intervention_status_select-button > span.ui-selectmenu-text"
dropdown_btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, CSS_DROPDOWN_BTN)))
dropdown_btn.click()

CSS_OPTION_TEMPORAIRE = "#ui-id-7"
option_TEMPORAIRE = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, CSS_OPTION_TEMPORAIRE)))
option_TEMPORAIRE.click()
time.sleep(2)

# --- 5. ATTENTE DU TABLEAU ---
wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#tbl_inter_pending")))

# --- 6. PRÉPARATION EXCEL AVEC PERSISTENCE ---
file_name = "resultats_interventions.xlsx"
decodeurs_vus = set()

if os.path.exists(file_name):
    wb = load_workbook(file_name)
    ws = wb.active
    # Charger les décodeurs déjà enregistrés
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[3]:
            for val in str(row[3]).split(", "):
                decodeurs_vus.add(val.strip())
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Résultats"
    ws.append(["Page", "Ligne", "ID Tech", "Décodeurs"])  # en-têtes

page_num = 1

# --- 7. BOUCLE PAGINATION ---
while True:
    try:
        print(f"--- Page {page_num} ---")
        wait.until(EC.presence_of_element_located((By.ID, "tbl_inter_pending")))

        voir_btns = driver.find_elements(By.XPATH, '//table[@id="tbl_inter_pending"]//a[contains(text(), "Voir")]')
        print(f"Nombre de lignes trouvées avec 'Voir' : {len(voir_btns)}")

        for i, btn in enumerate(voir_btns, start=1):
            try:
                driver.execute_script("arguments[0].click();", btn)
                wait.until(lambda d: len(d.window_handles) > 1)
                driver.switch_to.window(driver.window_handles[-1])

                wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "h3.ui-accordion-header")))
                headers = driver.find_elements(By.CSS_SELECTOR, "h3.ui-accordion-header")

                header_rdv = None
                header_cr = None
                for h in headers:
                    t = h.text.strip().lower()
                    if "rdv" in t:
                        header_rdv = h
                    if "compte" in t or "rendu" in t:
                        header_cr = h

                # --- EXTRAIRE ID TECH ---
                id_rdv = "ID_NON_TROUVÉ"
                if header_rdv:
                    driver.execute_script("arguments[0].click();", header_rdv)
                    time.sleep(1)
                    id_element = driver.find_element(
                        By.XPATH, "//*[contains(text(),'ID Tech')]/following::div[@class='accordion_input_div'][1]"
                    )
                    id_rdv = id_element.text.strip()
                    driver.execute_script("arguments[0].click();", header_rdv)

                # --- EXTRAIRE DÉCODEURS UNIQUES À 14 CHIFFRES ---
                decodeurs_uniques = []
                if header_cr:
                    driver.execute_script("arguments[0].click();", header_cr)
                    time.sleep(1)
                    decodeur_elements = driver.find_elements(By.XPATH, "//input[starts-with(@id,'ref_decodeur')]")
                    for el in decodeur_elements:
                        val = el.get_attribute("value").strip()
                        if re.fullmatch(r"\d{14}", val) and val not in decodeurs_vus:
                            decodeurs_uniques.append(val)
                            decodeurs_vus.add(val)
                    driver.execute_script("arguments[0].click();", header_cr)

                # --- ENREGISTRER DANS EXCEL ---
                if decodeurs_uniques:
                    ws.append([page_num, i, id_rdv, ", ".join(decodeurs_uniques)])
                    print(f"[Page {page_num} - Ligne {i}] ID Tech : {id_rdv} --- Nouveaux décodeurs (14 chiffres) : {', '.join(decodeurs_uniques)}")
                else:
                    print(f"[Page {page_num} - Ligne {i}] Aucun nouveau décodeur à 14 chiffres.")

                driver.close()
                driver.switch_to.window(driver.window_handles[0])

            except Exception as e:
                print(f"[Page {page_num} - Ligne {i}] Erreur :", e)

        # --- PASSER À LA PAGE SUIVANTE ---
        try:
            next_btn = driver.find_element(By.XPATH, "//a[contains(text(),'Suivant')]")
            if "ui-state-disabled" in next_btn.get_attribute("class"):
                print("Dernière page atteinte.")
                break
            else:
                driver.execute_script("arguments[0].click();", next_btn)
                page_num += 1
                time.sleep(2)
        except NoSuchElementException:
            print("Bouton 'Suivant' introuvable -> fin de la pagination.")
            break

    except Exception as e:
        print("Erreur sur la page :", e)
        break

# --- 8. SAUVEGARDE AVEC ÉCRASEMENT FORCÉ ---

def force_save(wb, file_path):
    try:
        wb.save(file_path)
        print(f"✔ Fichier enregistré : {file_path}")

    except PermissionError:
        # Le fichier est ouvert → on supprime puis on sauvegarde proprement
        print("❗ Le fichier est ouvert → suppression puis écrasement.")
        try:
            os.remove(file_path)  # suppression forcée
            wb.save(file_path)    # enregistrement propre
            print(f"✔ Fichier écrasé et mis à jour : {file_path}")
        except Exception as e:
            print(f"❌ Impossible d'écraser le fichier : {e}")

    except Exception as e:
        print(f"❌ Erreur lors de l'enregistrement : {e}")


force_save(wb, file_name)